import json
from pathlib import Path
from datetime import datetime, timezone, timedelta

DATA_DIR = Path(__file__).parent / "data"
HISTORY_FILE = DATA_DIR / "history.json"


def load_history():
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []


def save_history(records: list):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_FILE.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")


def append_signal(record: dict):
    history = load_history()
    history.append(record)
    save_history(history)


def export_xlsx(path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    records = load_history()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sinais"

    headers = ["Data/Hora", "Sinal", "Ativo", "Preço", "Timestamp"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F1F2E")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(records, 2):
        ws.cell(row=i, column=1, value=r.get("time_str", ""))
        ws.cell(row=i, column=2, value=r.get("action", ""))
        ws.cell(row=i, column=3, value=r.get("ticker", ""))
        ws.cell(row=i, column=4, value=r.get("price", ""))
        ws.cell(row=i, column=5, value=r.get("time", ""))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18

    wb.save(path)
    return len(records)


def _fmt_time_br(time_ms):
    try:
        t = datetime.fromtimestamp(time_ms / 1000, tz=timezone(timedelta(hours=-3)))
        return t.strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return str(time_ms)
